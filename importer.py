import openpyxl
def import_from_excel(path: str, store):
wb = openpyxl.load_workbook(path)
ws = wb.active
header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
idx = {name:i for i,name in enumerate(header)}
n = 0
for row in ws.iter_rows(min_row=2, values_only=True):
if not row[idx["question"]]: continue
q = {
"type": row[idx["type"]] or "qa",
"category": row[idx["category"]] or "测试理论",
"difficulty": row[idx["difficulty"]] or "初",
"question": row[idx["question"]],
"options": (row[idx["options"]].split("||") if row[idx["options"]] else []),
"answer": str(row[idx["answer"]] if row[idx["answer"]] is not None else ""),
"tags": [t.strip() for t in (row[idx["tags"]] or "").split(",") if t.strip()]
}
store.add_question(q)
n += 1
return n
